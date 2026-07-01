from dotenv import load_dotenv
from pathlib import Path
from contextlib import asynccontextmanager

import io
import os
import uuid
import logging

from datetime import datetime, timezone, timedelta
from typing import Optional, Any

import bcrypt
import jwt
import openpyxl

from motor.motor_asyncio import AsyncIOMotorClient

from fastapi import (
    FastAPI,
    APIRouter,
    HTTPException,
    Depends,
    Request,
)

from fastapi.responses import StreamingResponse

from starlette.middleware.cors import CORSMiddleware

from pydantic import BaseModel, EmailStr, Field

from openpyxl.styles import (
    Font,
    PatternFill,
    Alignment,
    Border,
    Side,
)
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter

# =========================================================
# ENV
# =========================================================

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / ".env", override=True)

MONGO_URL = os.getenv("MONGO_URL")
DB_NAME = os.getenv("DB_NAME", "bayer_db")
JWT_SECRET = os.getenv("JWT_SECRET")

if not MONGO_URL:
    raise RuntimeError("MONGO_URL não configurado")

if not JWT_SECRET:
    raise RuntimeError("JWT_SECRET não configurado — defina a variável de ambiente JWT_SECRET")

# =========================================================
# DATABASE
# =========================================================

client = AsyncIOMotorClient(MONGO_URL)
db = client[DB_NAME]

# =========================================================
# LOGGING
# =========================================================

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s",
)

logger = logging.getLogger(__name__)

# =========================================================
# CONSTANTS
# =========================================================

JWT_ALGORITHM = "HS256"
ACCESS_TOKEN_EXPIRE_MINUTES = 60 * 24 * 7

DEFAULT_PRODUCTS = [
  { "name": "ALSYSTIN",     "abbr": "ALS", "category": "Inseticida",               "subcategory": "Regulador de Crescimento de Insetos" },
  { "name": "BULLDOCK",     "abbr": "BUL", "category": "Inseticida",               "subcategory": "Piretróide" },
  { "name": "CONNECT",      "abbr": "CON", "category": "Inseticida",               "subcategory": "Neonicotinoide + Piretróide" },
  { "name": "CURBIX",       "abbr": "CUR", "category": "Inseticida",               "subcategory": "Sulfoximina" },
  { "name": "FOX XPRO",     "abbr": "FXX", "category": "Fungicida",                "subcategory": "QoI + DMI + SDHI" },
  { "name": "FOX ULTRA",    "abbr": "FUL", "category": "Fungicida",                "subcategory": "QoI + DMI + Morfolina" },
  { "name": "NATIVO",       "abbr": "NAT", "category": "Fungicida",                "subcategory": "Triazol + QoI" },
  { "name": "OBERON",       "abbr": "OBE", "category": "Acaricida",                "subcategory": "Cetoenol" },
  { "name": "PREMIER PLUS", "abbr": "PRP", "category": "Inseticida",               "subcategory": "Neonicotinoide + Piretróide" },
  { "name": "PROVADO",      "abbr": "PRO", "category": "Inseticida",               "subcategory": "Neonicotinoide" },
  { "name": "SPHERE MAX",   "abbr": "SPM", "category": "Fungicida",                "subcategory": "QoI + Triazol" },
  { "name": "FINISH",       "abbr": "FIN", "category": "Regulador de Crescimento", "subcategory": "Etefon" },
  { "name": "SOBERAN",      "abbr": "SOB", "category": "Fungicida",                "subcategory": "SDHI + QoI" },
  { "name": "VERANGO",      "abbr": "VER", "category": "Fungicida",                "subcategory": "SDHI + QoI" },
  { "name": "GAUCHO",       "abbr": "GAU", "category": "Inseticida",               "subcategory": "Neonicotinoide – Tratamento de Sementes" },
]

# =========================================================
# HELPERS
# =========================================================


def hash_password(password: str) -> str:
    return bcrypt.hashpw(
        password.encode("utf-8"),
        bcrypt.gensalt(),
    ).decode("utf-8")


def verify_password(
    plain: str,
    hashed: str,
) -> bool:
    return bcrypt.checkpw(
        plain.encode("utf-8"),
        hashed.encode("utf-8"),
    )


def create_access_token(
    user_id: str,
    email: str,
) -> str:
    expire = datetime.now(
        timezone.utc
    ) + timedelta(
        minutes=ACCESS_TOKEN_EXPIRE_MINUTES
    )

    payload = {
        "sub": user_id,
        "email": email,
        "type": "access",
        "exp": int(expire.timestamp()),
    }

    return jwt.encode(
        payload,
        JWT_SECRET,
        algorithm=JWT_ALGORITHM,
    )


def create_reset_token(
    user_id: str,
    email: str,
) -> str:
    expire = datetime.now(
        timezone.utc
    ) + timedelta(
        minutes=30
    )

    payload = {
        "sub": user_id,
        "email": email,
        "type": "reset",
        "exp": int(expire.timestamp()),
    }

    return jwt.encode(
        payload,
        JWT_SECRET,
        algorithm=JWT_ALGORITHM,
    )


def auto_abbreviate(name: str) -> str:
    name = (name or "").strip()

    if not name:
        return ""

    for p in DEFAULT_PRODUCTS:
        if p["name"].lower() == name.lower():
            return p["abbr"]

    cleaned = "".join(
        c for c in name if c.isalpha()
    )

    if cleaned:
        return cleaned[:3].upper()

    return name[:3].upper()


def greeting_for_now() -> str:
    now = datetime.now(
        timezone.utc
    ) - timedelta(hours=3)

    hour = now.hour

    if 5 <= hour < 12:
        return "Bom dia"

    if 12 <= hour < 18:
        return "Boa tarde"

    return "Boa noite"


# =========================================================
# MODELS
# =========================================================


class UserPublic(BaseModel):
    id: str
    email: str
    name: str
    role: str = "user"
    matricula: Optional[str] = None
    department: Optional[str] = None


class RegisterRequest(BaseModel):
    email: EmailStr
    password: str = Field(min_length=6)
    name: str = Field(min_length=1)
    matricula: Optional[str] = None


class LoginRequest(BaseModel):
    identifier: str
    password: str


class TokenResponse(BaseModel):
    access_token: str
    token_type: str = "bearer"
    user: UserPublic


class ForgotPasswordRequest(BaseModel):
    identifier: str


class ResetPasswordRequest(BaseModel):
    token: str
    password: str = Field(min_length=6)


class PasswordResetResponse(BaseModel):
    message: str = "Senha redefinida com sucesso"


class ProductionItem(BaseModel):
    id: str = Field(
        default_factory=lambda: str(uuid.uuid4())
    )

    date: str
    unit: str
    sc: str

    product: str
    product_abbr: str = ""

    batch: str

    quantity: Optional[float] = None
    quantity_unit: str = "bag"

    material_status: str = "Disponível"
    situation: str = "A preparar"

    observation: str = ""

    created_at: str = Field(
        default_factory=lambda:
        datetime.now(timezone.utc).isoformat()
    )

    updated_at: str = Field(
        default_factory=lambda:
        datetime.now(timezone.utc).isoformat()
    )


class ProductionItemCreate(BaseModel):
    date: str
    unit: str
    sc: str

    product: str
    batch: str

    quantity: Optional[float] = None
    quantity_unit: str = "kg"

    material_status: str = "Disponível"
    situation: str = "A preparar"

    observation: str = ""


class ProductionItemUpdate(BaseModel):
    unit: Optional[str] = None
    sc: Optional[str] = None

    product: Optional[str] = None
    batch: Optional[str] = None

    quantity: Optional[float] = None
    quantity_unit: Optional[str] = None

    material_status: Optional[str] = None
    situation: Optional[str] = None

    observation: Optional[str] = None
    date: Optional[str] = None


class ProductCreate(BaseModel):
    name: str
    abbr: Optional[str] = None


class ReportRequest(BaseModel):
    date: str
    extra_observations: Optional[str] = None


class Recipe(BaseModel):
    name: str
    abbr: str
    description: Optional[str] = None
    ingredients: Optional[list[str]] = None
    procedure: Optional[str] = None
    category: str = "produtos"


class GalleryImage(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    user_id: str
    url: str
    caption: Optional[str] = None
    category: str = "geral"
    created_at: str = Field(
        default_factory=lambda:
        datetime.now(timezone.utc).isoformat()
    )


# =========================================================
# APP
# =========================================================


@asynccontextmanager
async def lifespan(app: FastAPI):

    await db.users.create_index(
        "email",
        unique=True,
    )

    await db.users.create_index(
        "id",
        unique=True,
    )

    await db.production_items.create_index(
        "id",
        unique=True,
    )

    await db.production_items.create_index(
        "date"
    )

    await db.products.create_index(
        "name",
        unique=True,
    )

    await db.password_resets.create_index(
        "user_id",
        unique=True,
    )

    await db.gallery.create_index(
        "id",
        unique=True,
    )

    await db.gallery.create_index(
        "created_at"
    )

    admin_email = os.getenv(
        "ADMIN_EMAIL",
        "admin@bayer.com",
    ).lower()

    admin_password = os.getenv(
        "ADMIN_PASSWORD",
        "admin123",
    )

    existing = await db.users.find_one({
        "email": admin_email
    })

    if existing is None:

        await db.users.insert_one({
            "id": str(uuid.uuid4()),
            "email": admin_email,
            "password_hash": hash_password(
                admin_password
            ),
            "name": "Administrador",
            "role": "admin",
            "created_at": datetime.now(
                timezone.utc
            ).isoformat(),
        })

        logger.info(
            "Admin criado: %s",
            admin_email,
        )

    yield

    client.close()


app = FastAPI(
    title="Bayer Production Control",
    lifespan=lifespan,
)

api_router = APIRouter(
    prefix="/api"
)

# =========================================================
# CORS
# =========================================================

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.getenv(
        "CORS_ORIGINS",
        "*",
    ).split(","),
    allow_methods=["*"],
    allow_headers=["*"],
)

# =========================================================
# AUTH
# =========================================================


async def get_current_user(
    request: Request,
) -> dict[str, Any]:

    auth_header = request.headers.get(
        "Authorization",
        "",
    )

    token = ""

    if auth_header.startswith("Bearer "):
        token = auth_header[7:]

    if not token:
        raise HTTPException(
            status_code=401,
            detail="Não autenticado",
        )

    try:

        payload = jwt.decode(
            token,
            JWT_SECRET,
            algorithms=[JWT_ALGORITHM],
        )

        if payload.get("type") != "access":
            raise HTTPException(
                status_code=401,
                detail="Token inválido",
            )

    except jwt.ExpiredSignatureError:
        raise HTTPException(
            status_code=401,
            detail="Token expirado",
        )

    except jwt.InvalidTokenError:
        raise HTTPException(
            status_code=401,
            detail="Token inválido",
        )

    user = await db.users.find_one(
        {"id": payload["sub"]},
        {"_id": 0, "password_hash": 0},
    )

    if not user:
        raise HTTPException(
            status_code=401,
            detail="Usuário não encontrado",
        )

    return user


# =========================================================
# ROUTES
# =========================================================

# Health check público (sem autenticação) - DEVE estar ANTES da API
@app.get("/api/health")
async def health_public():
    """
    Health check endpoint público para monitoramento do app.
    Não requer autenticação.
    """
    try:
        # Verificar se banco de dados está conectado
        await db.command("ping")

        return {
            "status": "healthy",
            "service": "Bayer Production Control",
            "timestamp": datetime.now(timezone.utc).isoformat(),
            "version": "2.0.0",
            "database": "connected",
        }
    except Exception as e:
        logger.error("Health check falhou: %s", str(e))
        return {
            "status": "degraded",
            "service": "Bayer Production Control",
            "timestamp": datetime.now(timezone.utc).isoformat(),
            "version": "2.0.0",
            "database": "disconnected",
        }


@api_router.get("/")
async def root():
    return {
        "service": "Bayer Production Control",
        "ok": True,
        "version": "2.0.0",
    }


@api_router.get("/health")
async def health():
    """Health check endpoint para monitoramento 24/7"""
    return {
        "status": "healthy",
        "service": "Bayer Production Control",
        "timestamp": datetime.now(timezone.utc).isoformat(),
        "version": "2.0.0",
        "database": "connected",
    }


# =========================================================
# AUTH ROUTES
# =========================================================


@api_router.post(
    "/auth/register",
    response_model=TokenResponse,
)
async def register(
    payload: RegisterRequest
):

    email = payload.email.lower()

    existing = await db.users.find_one({
        "email": email
    })

    if existing:
        raise HTTPException(
            status_code=400,
            detail="E-mail já cadastrado",
        )

    if payload.matricula:
        mat_existing = await db.users.find_one({
            "matricula": payload.matricula.strip()
        })
        if mat_existing:
            raise HTTPException(
                status_code=400,
                detail="Matrícula já cadastrada",
            )

    user_id = str(uuid.uuid4())

    await db.users.insert_one({
        "id": user_id,
        "email": email,
        "password_hash": hash_password(
            payload.password
        ),
        "name": payload.name,
        "role": "user",
        "matricula": payload.matricula.strip() if payload.matricula else None,
        "department": None,
        "created_at": datetime.now(
            timezone.utc
        ).isoformat(),
    })

    token = create_access_token(
        user_id,
        email,
    )

    return TokenResponse(
        access_token=token,
        user=UserPublic(
            id=user_id,
            email=email,
            name=payload.name,
            role="user",
            matricula=payload.matricula.strip() if payload.matricula else None,
            department=None,
        ),
    )


@api_router.post(
    "/auth/login",
    response_model=TokenResponse,
)
async def login(
    payload: LoginRequest
):
    identifier = payload.identifier.strip()

    # Try by email first, then by matricula
    user = await db.users.find_one({"email": identifier.lower()})
    if not user:
        user = await db.users.find_one({"matricula": identifier})

    if not user:
        raise HTTPException(
            status_code=401,
            detail="Credenciais inválidas",
        )

    if not verify_password(
        payload.password,
        user["password_hash"],
    ):
        raise HTTPException(
            status_code=401,
            detail="Credenciais inválidas",
        )

    token = create_access_token(
        user["id"],
        user["email"],
    )

    return TokenResponse(
        access_token=token,
        user=UserPublic(
            id=user["id"],
            email=user["email"],
            name=user.get("name", ""),
            role=user.get("role", "user"),
            matricula=user.get("matricula"),
            department=user.get("department"),
        ),
    )


@api_router.get(
    "/auth/me",
    response_model=UserPublic,
)
async def me(
    user: dict = Depends(
        get_current_user
    )
):

    return UserPublic(
        id=user["id"],
        email=user["email"],
        name=user.get("name", ""),
        role=user.get("role", "user"),
        matricula=user.get("matricula"),
        department=user.get("department"),
    )


class UpdateProfileRequest(BaseModel):
    department: Optional[str] = None
    name: Optional[str] = None


@api_router.patch(
    "/auth/me",
    response_model=UserPublic,
)
async def update_me(
    payload: UpdateProfileRequest,
    user: dict = Depends(get_current_user),
):
    update_fields: dict = {}
    if payload.department is not None:
        update_fields["department"] = payload.department
    if payload.name is not None:
        update_fields["name"] = payload.name

    if update_fields:
        await db.users.update_one(
            {"id": user["id"]},
            {"$set": update_fields},
        )

    updated = await db.users.find_one(
        {"id": user["id"]},
        {"_id": 0, "password_hash": 0},
    )

    return UserPublic(
        id=updated["id"],
        email=updated["email"],
        name=updated.get("name", ""),
        role=updated.get("role", "user"),
        matricula=updated.get("matricula"),
        department=updated.get("department"),
    )


@api_router.post("/auth/forgot-password")
async def forgot_password(
    payload: ForgotPasswordRequest,
):

    identifier = payload.identifier.strip().lower()

    user = await db.users.find_one({
        "$or": [
            {"email": identifier},
            {"matricula": identifier},
        ]
    })

    if not user:
        return {
            "found": False,
            "message": "Usuário não encontrado",
        }

    email = user["email"]
    reset_token = create_reset_token(
        user["id"],
        email,
    )

    await db.password_resets.update_one(
        {"user_id": user["id"]},
        {
            "$set": {
                "user_id": user["id"],
                "email": email,
                "token": reset_token,
                "created_at": datetime.now(
                    timezone.utc
                ).isoformat(),
                "used": False,
            }
        },
        upsert=True,
    )

    logger.info(
        "Reset token criado para: %s",
        email,
    )

    return {
        "found": True,
        "token": reset_token,
        "message": "Usuário verificado. Defina sua nova senha.",
    }


@api_router.post(
    "/auth/reset-password",
    response_model=PasswordResetResponse,
)
async def reset_password(
    payload: ResetPasswordRequest,
):

    try:

        decoded = jwt.decode(
            payload.token,
            JWT_SECRET,
            algorithms=[JWT_ALGORITHM],
        )

        if decoded.get("type") != "reset":
            raise HTTPException(
                status_code=400,
                detail="Token inválido",
            )

    except jwt.ExpiredSignatureError:
        raise HTTPException(
            status_code=400,
            detail="Link expirado. Solicite um novo reset",
        )

    except jwt.InvalidTokenError:
        raise HTTPException(
            status_code=400,
            detail="Token inválido",
        )

    user_id = decoded.get("sub")

    user = await db.users.find_one({
        "id": user_id
    })

    if not user:
        raise HTTPException(
            status_code=404,
            detail="Usuário não encontrado",
        )

    reset_record = await db.password_resets.find_one({
        "user_id": user_id,
    })

    if reset_record and reset_record.get("used"):
        raise HTTPException(
            status_code=400,
            detail="Este link de reset já foi utilizado",
        )

    await db.users.update_one(
        {"id": user_id},
        {
            "$set": {
                "password_hash": hash_password(
                    payload.password
                ),
                "updated_at": datetime.now(
                    timezone.utc
                ).isoformat(),
            }
        },
    )

    await db.password_resets.update_one(
        {"user_id": user_id},
        {
            "$set": {
                "used": True,
                "used_at": datetime.now(
                    timezone.utc
                ).isoformat(),
            }
        },
    )

    logger.info(
        "Senha resetada para: %s",
        user["email"],
    )

    return PasswordResetResponse()


# =========================================================
# ITEMS
# =========================================================


@api_router.get("/items")
async def list_items(
    date: Optional[str] = None,
    user: dict = Depends(get_current_user),
):

    query = {}

    if date:
        query["date"] = date

    items = await db.production_items.find(
        query,
        {"_id": 0},
    ).sort(
        "created_at",
        1,
    ).to_list(2000)

    return items


@api_router.post(
    "/items",
    response_model=ProductionItem,
)
async def create_item(
    payload: ProductionItemCreate,
    user: dict = Depends(get_current_user),
):

    item = ProductionItem(
        **payload.model_dump(),
        product_abbr=auto_abbreviate(
            payload.product
        ),
    )

    await db.production_items.insert_one(
        item.model_dump()
    )

    return item


@api_router.put(
    "/items/{item_id}",
    response_model=ProductionItem,
)
async def update_item(
    item_id: str,
    payload: ProductionItemUpdate,
    user: dict = Depends(get_current_user),
):

    existing = await db.production_items.find_one(
        {"id": item_id},
        {"_id": 0},
    )

    if not existing:
        raise HTTPException(
            status_code=404,
            detail="Item não encontrado",
        )

    update = {
        k: v
        for k, v in payload.model_dump(
            exclude_unset=True
        ).items()
        if v is not None
    }

    if "product" in update:
        update["product_abbr"] = auto_abbreviate(
            update["product"]
        )

    update["updated_at"] = datetime.now(
        timezone.utc
    ).isoformat()

    await db.production_items.update_one(
        {"id": item_id},
        {"$set": update},
    )

    merged = {
        **existing,
        **update,
    }

    return ProductionItem(**merged)


@api_router.delete("/items/{item_id}")
async def delete_item(
    item_id: str,
    user: dict = Depends(get_current_user),
):

    result = await db.production_items.delete_one({
        "id": item_id
    })

    if result.deleted_count == 0:
        raise HTTPException(
            status_code=404,
            detail="Item não encontrado",
        )

    return {"ok": True}


# =========================================================
# PRODUCTS
# =========================================================


@api_router.get("/products")
async def list_products(
    user: dict = Depends(get_current_user),
):

    custom = await db.products.find(
        {},
        {"_id": 0},
    ).to_list(500)

    merged = {
        p["name"]: p["abbr"]
        for p in DEFAULT_PRODUCTS
    }

    for item in custom:
        merged[item["name"]] = item["abbr"]

    return [
        {
            "name": k,
            "abbr": v,
        }
        for k, v in merged.items()
    ]


@api_router.post("/products")
async def add_product(
    payload: ProductCreate,
    user: dict = Depends(get_current_user),
):

    abbr = (
        payload.abbr
        or auto_abbreviate(payload.name)
    ).upper()

    await db.products.update_one(
        {"name": payload.name},
        {
            "$set": {
                "name": payload.name,
                "abbr": abbr,
            }
        },
        upsert=True,
    )

    return {
        "name": payload.name,
        "abbr": abbr,
    }


# =========================================================
# RECIPES
# =========================================================

DEFAULT_RECIPES = [
    {
        "name": "VERANGO",
        "abbr": "VER",
        "category": "produtos",
        "description": "Fungicida sistêmico para folhas",
        "ingredients": ["Trifloxystrobin", "Fluopyram"],
        "procedure": "Aplicar de acordo com recomendações técnicas",
    },
    {
        "name": "NATIVO",
        "abbr": "NAT",
        "category": "produtos",
        "description": "Fungicida de contato e sistêmico",
        "ingredients": ["Trifloxystrobin", "Tebucconazole"],
        "procedure": "Pulverizar uniformemente a cultura",
    },
    {
        "name": "OBERON",
        "abbr": "OBE",
        "category": "produtos",
        "description": "Acaricida seletivo",
        "ingredients": ["Spiromesifen"],
        "procedure": "Indicado para controle de ácaros",
    },
    {
        "name": "FOX XPRO",
        "abbr": "FXX",
        "category": "produtos",
        "description": "Fungicida tríplice ação",
        "ingredients": ["Trifloxystrobin", "Protioconazole", "Bixafen"],
        "procedure": "Aplicação em estágio inicial de infecção",
    },
    {
        "name": "FOX",
        "abbr": "FOX",
        "category": "produtos",
        "description": "Fungicida dupla ação",
        "ingredients": ["Trifloxystrobin", "Protioconazole"],
        "procedure": "Pulverização recomendada",
    },
    {
        "name": "BELT",
        "abbr": "BEL",
        "category": "produtos",
        "description": "Inseticida neonicotinóide",
        "ingredients": ["Flubendiamide"],
        "procedure": "Controle de lepidópteros",
    },
]


@api_router.get("/recipes")
async def list_recipes(
    category: Optional[str] = None,
    user: dict = Depends(get_current_user),
):

    recipes = DEFAULT_RECIPES

    if category:
        recipes = [
            r for r in recipes
            if r.get("category") == category
        ]

    return recipes


# =========================================================
# GALLERY
# =========================================================


@api_router.get("/gallery")
async def list_gallery(
    category: Optional[str] = None,
    skip: int = 0,
    limit: int = 50,
    user: dict = Depends(get_current_user),
):

    query = {}

    if category:
        query["category"] = category

    images = await db.gallery.find(
        query,
        {"_id": 0},
    ).sort(
        "created_at",
        -1,
    ).skip(skip).limit(limit).to_list(limit)

    return images


@api_router.post("/gallery")
async def add_gallery_image(
    payload: dict,
    user: dict = Depends(get_current_user),
):

    image_doc = {
        "id": str(uuid.uuid4()),
        "user_id": user["id"],
        "user_name": user.get("name", ""),
        "url": payload.get("url"),
        "caption": payload.get("caption", ""),
        "category": payload.get("category", "geral"),
        "created_at": datetime.now(
            timezone.utc
        ).isoformat(),
    }

    await db.gallery.insert_one(image_doc)

    return image_doc


@api_router.delete("/gallery/{image_id}")
async def delete_gallery_image(
    image_id: str,
    user: dict = Depends(get_current_user),
):

    img = await db.gallery.find_one({
        "id": image_id
    })

    if not img:
        raise HTTPException(
            status_code=404,
            detail="Imagem não encontrada",
        )

    if img["user_id"] != user["id"] and user.get("role") != "admin":
        raise HTTPException(
            status_code=403,
            detail="Não autorizado",
        )

    await db.gallery.delete_one({
        "id": image_id
    })

    return {"ok": True}


# =========================================================
# REPORT
# =========================================================


def build_report(
    items: list[dict],
    greeting: str,
    extra_obs: Optional[str],
):

    lines = []

    lines.append(
        f"*{greeting}, segue a situação dos materiais para o próximo turno:*"
    )

    lines.append("")

    for item in items:

        qty = ""

        if item.get("quantity") is not None:
            qty = (
                f" - {item['quantity']:g}"
                f"{item.get('quantity_unit', '')}"
            )

        obs = item.get(
            "observation",
            "",
        ).strip()

        obs_text = f" — _{obs}_" if obs else ""

        lines.append(
            f"• {item.get('unit')} | "
            f"{item.get('sc')} | "
            f"{item.get('product')} | "
            f"Lote {item.get('batch')}"
            f"{qty} | "
            f"{item.get('situation')}"
            f"{obs_text}"
        )

    if extra_obs:
        lines.append("")
        lines.append(f"📝 {extra_obs}")

    return "\n".join(lines)


@api_router.post("/reports/whatsapp")
async def whatsapp_report(
    payload: ReportRequest,
    user: dict = Depends(get_current_user),
):

    items = await db.production_items.find(
        {"date": payload.date},
        {"_id": 0},
    ).to_list(2000)

    greeting = greeting_for_now()

    text = build_report(
        items,
        greeting,
        payload.extra_observations,
    )

    return {
        "text": text,
        "count": len(items),
        "greeting": greeting,
    }


# =========================================================
# EXCEL EXPORT  (styled, with Bayer branding)
# =========================================================

LOGO_PATH = Path(__file__).parent / "bayer_logo.png"

# colour palette
C_GREEN_DARK  = "0A4A20"
C_GREEN_MED   = "00A04E"
C_GREEN_LIGHT = "E8F5E9"
C_GREEN_ALT   = "F0FAF4"
C_WHITE       = "FFFFFF"
C_GREY_TEXT   = "444444"
C_BORDER      = "BDBDBD"
C_HEADER_TXT  = "FFFFFF"

def _thin_border(color=C_BORDER):
    s = Side(border_style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def _apply_header_style(cell, font_size=10):
    cell.fill   = PatternFill("solid", fgColor=C_GREEN_MED)
    cell.font   = Font(bold=True, color=C_HEADER_TXT, size=font_size, name="Calibri")
    cell.border = _thin_border()
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

def _apply_data_style(cell, alt_row=False, align="left"):
    cell.fill      = PatternFill("solid", fgColor=C_GREEN_ALT if alt_row else C_WHITE)
    cell.font      = Font(size=10, color=C_GREY_TEXT, name="Calibri")
    cell.border    = _thin_border("E0E0E0")
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)


@api_router.get("/export/excel")
async def export_excel(
    date: str,
    user: dict = Depends(get_current_user),
):
    items = await db.production_items.find(
        {"date": date}, {"_id": 0}
    ).to_list(2000)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Produção {date}"

    # ── row heights
    ws.row_dimensions[1].height = 70   # logo row
    ws.row_dimensions[2].height = 22   # title
    ws.row_dimensions[3].height = 16   # subtitle
    ws.row_dimensions[4].height = 10   # spacer
    ws.row_dimensions[5].height = 28   # column headers
    ws.sheet_view.showGridLines = False

    HEADERS = [
        "Unidade", "SC", "Produto", "Abrev.",
        "Lote", "Quantidade", "Status MP", "Situação", "Observação",
    ]
    NUM_COLS = len(HEADERS)

    # ── merge top rows for branding banner
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=NUM_COLS)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=NUM_COLS)
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=NUM_COLS)
    ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=NUM_COLS)

    # banner background
    banner_fill = PatternFill("solid", fgColor=C_GREEN_DARK)
    for row in range(1, 5):
        cell = ws.cell(row=row, column=1)
        cell.fill = banner_fill

    # title text
    title_cell = ws.cell(row=2, column=1)
    title_cell.value = "BAYER OPERACIONAL  ·  Controle de Produção"
    title_cell.font  = Font(bold=True, size=16, color=C_WHITE, name="Calibri")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    # subtitle
    sub_cell = ws.cell(row=3, column=1)
    sub_cell.value = f"Data: {date}   |   Exportado por: {user.get('name', 'Operador')}   |   Bayer S.A. — Produção Industrial"
    sub_cell.font  = Font(size=9, color="AAFFAA", italic=True, name="Calibri")
    sub_cell.alignment = Alignment(horizontal="center", vertical="center")

    # logo
    if LOGO_PATH.exists():
        try:
            img = XLImage(str(LOGO_PATH))
            img.width  = 58
            img.height = 58
            img.anchor = "A1"
            ws.add_image(img)
        except Exception:
            pass

    # ── column headers (row 5)
    for col_idx, h in enumerate(HEADERS, 1):
        cell = ws.cell(row=5, column=col_idx, value=h)
        _apply_header_style(cell)

    # ── data rows
    WIDTHS = [13, 6, 22, 9, 16, 13, 17, 18, 36]
    for i, w in enumerate(WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # status colours for "Situação" column (col 8)
    STATUS_FILLS = {
        "Preparado":  "C8E6C9",
        "Pendente":   "FFF9C4",
        "Cancelado":  "FFCDD2",
        "Em andamento": "BBDEFB",
    }

    for idx, item in enumerate(items):
        row_num = 6 + idx
        alt = (idx % 2 == 1)
        ws.row_dimensions[row_num].height = 18

        qty = ""
        if item.get("quantity") is not None:
            qty = f"{item['quantity']:g} {item.get('quantity_unit', '')}".strip()

        values = [
            item.get("unit", ""),
            item.get("sc", ""),
            item.get("product", ""),
            item.get("product_abbr", ""),
            item.get("batch", ""),
            qty,
            item.get("material_status", ""),
            item.get("situation", ""),
            item.get("observation", ""),
        ]
        aligns = ["center","center","left","center","center","center","center","center","left"]

        for col_idx, (val, aln) in enumerate(zip(values, aligns), 1):
            cell = ws.cell(row=row_num, column=col_idx, value=val)
            _apply_data_style(cell, alt_row=alt, align=aln)

        # colour-code situação
        sit = item.get("situation", "")
        for key, fill_color in STATUS_FILLS.items():
            if key.lower() in str(sit).lower():
                ws.cell(row=row_num, column=8).fill = PatternFill("solid", fgColor=fill_color)
                ws.cell(row=row_num, column=8).font = Font(size=10, color="333333", bold=True, name="Calibri")
                break

    # ── auto-filter on header row
    ws.auto_filter.ref = f"A5:{get_column_letter(NUM_COLS)}5"

    # ── freeze panes below header
    ws.freeze_panes = "A6"

    # ── footer note in last row + 2
    last_row = 6 + len(items)
    ws.merge_cells(start_row=last_row, start_column=1, end_row=last_row, end_column=NUM_COLS)
    footer = ws.cell(row=last_row, column=1)
    footer.value = f"Bayer S.A. · Documento gerado automaticamente pelo sistema Bayer Operacional em {datetime.now(timezone.utc).strftime('%d/%m/%Y %H:%M')} UTC"
    footer.font  = Font(size=8, color="AAAAAA", italic=True, name="Calibri")
    footer.alignment = Alignment(horizontal="right")

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"bayer_planilha_{date}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# =========================================================
# PASSAGEM DE SERVIÇO
# =========================================================

class LotEntry(BaseModel):
    unit: str = ""
    sc: str = ""
    product: str = ""
    lot: str = ""
    status: str = ""
    notes: str = ""

class ReceiptEntry(BaseModel):
    product: str = ""
    qty: str = ""

class HandoverCreate(BaseModel):
    shift: str
    lots: list[LotEntry] = []
    receipts: list[ReceiptEntry] = []
    observations: str = ""
    participants: list[str] = []

@api_router.post("/handover")
async def create_handover(
    body: HandoverCreate,
    user: dict = Depends(get_current_user),
):
    doc = {
        "shift": body.shift,
        "lots": [lot.model_dump() for lot in body.lots],
        "receipts": [r.model_dump() for r in body.receipts],
        "observations": body.observations,
        "participants": body.participants,
        "created_by": user.get("email", ""),
        "created_by_name": user.get("name", ""),
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    result = await db.handovers.insert_one(doc)
    doc["_id"] = str(result.inserted_id)
    return doc

@api_router.get("/handover")
async def list_handovers(
    shift: Optional[str] = None,
    limit: int = 40,
    user: dict = Depends(get_current_user),
):
    query: dict = {}
    if shift:
        query["shift"] = shift
    projection = {
        "_id": 1, "shift": 1, "lots": 1, "receipts": 1,
        "observations": 1, "participants": 1,
        "created_by": 1, "created_by_name": 1, "created_at": 1,
    }
    cursor = db.handovers.find(query, projection)
    cursor = cursor.sort("created_at", -1).limit(limit)
    docs = await cursor.to_list(limit)
    for d in docs:
        d["_id"] = str(d["_id"])
        if "receipts" not in d:
            d["receipts"] = []
    return docs


# =========================================================
# HANDOVER EDIT / DELETE
# =========================================================

HANDOVER_EDIT_WINDOW_MINUTES = 15

def _can_modify_handover(h_created_at: str, user: dict) -> bool:
    if user.get("role") == "admin":
        return True
    try:
        created = datetime.fromisoformat(h_created_at.replace("Z", "+00:00"))
        return datetime.now(timezone.utc) - created <= timedelta(minutes=HANDOVER_EDIT_WINDOW_MINUTES)
    except Exception:
        return False

@api_router.put("/handover/{handover_id}")
async def update_handover(
    handover_id: str,
    body: HandoverCreate,
    user: dict = Depends(get_current_user),
):
    from bson import ObjectId
    try:
        oid = ObjectId(handover_id)
    except Exception:
        raise HTTPException(status_code=400, detail="ID inválido")
    existing = await db.handovers.find_one({"_id": oid})
    if not existing:
        raise HTTPException(status_code=404, detail="Passagem não encontrada")
    if not _can_modify_handover(existing.get("created_at", ""), user):
        raise HTTPException(status_code=403, detail="Prazo de edição encerrado (15 minutos)")
    update_data = {
        "shift": body.shift,
        "lots": [lot.model_dump() for lot in body.lots],
        "receipts": [r.model_dump() for r in body.receipts],
        "observations": body.observations,
        "participants": body.participants,
        "updated_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.handovers.update_one({"_id": oid}, {"$set": update_data})
    updated = await db.handovers.find_one({"_id": oid})
    updated["_id"] = str(updated["_id"])
    return updated

@api_router.delete("/handover/{handover_id}")
async def delete_handover(
    handover_id: str,
    user: dict = Depends(get_current_user),
):
    from bson import ObjectId
    try:
        oid = ObjectId(handover_id)
    except Exception:
        raise HTTPException(status_code=400, detail="ID inválido")
    existing = await db.handovers.find_one({"_id": oid})
    if not existing:
        raise HTTPException(status_code=404, detail="Passagem não encontrada")
    if not _can_modify_handover(existing.get("created_at", ""), user):
        raise HTTPException(status_code=403, detail="Prazo de exclusão encerrado (15 minutos)")
    await db.handovers.delete_one({"_id": oid})
    return {"ok": True}




# =========================================================
# INCLUDE ROUTER
# =========================================================

app.include_router(api_router)
